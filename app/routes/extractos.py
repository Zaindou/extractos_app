from flask import render_template, jsonify, request, Response, Blueprint, current_app
from app.models import Cliente, Producto, Extracto
from app.utils.validators import (
    is_valid_string,
    is_valid_name,
    is_valid_email,
    is_valid_integer,
    is_valid_float,
    validate_row,
)
from flask_weasyprint import HTML, render_pdf
from app.utils.helpers import (
    period_to_date,
    allowed_file,
    get_primary_email,
    date_to_period,
    encrypt_pdf,
)
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
from datetime import datetime
from config import Config
from hashlib import sha256
from app import db, cache, logging
import requests
import json
import uuid
import os
import csv

extractos = Blueprint("extractos", __name__)


# @extractos.route("/test-template")
# def test_email_template():
#     return render_template("extracto_estado_cuenta.html", cliente=Cliente.query.first(), registros=Producto.query.all(), periodo="2021-01-01", generation_date=datetime.now().strftime("%d/%m/%Y"))


@extractos.route("/<tipo_extracto>/<periodo>/<id_contacto>.pdf")
@cache.cached(
    timeout=int(Config.CACHE_TIMEOUT),
    key_prefix=lambda: sha256(request.full_path.encode()).hexdigest(),
)
def extracto_pdf(periodo, id_contacto, tipo_extracto):
    try:
        formatted_period = period_to_date(periodo)
        generation_date = datetime.now().strftime("%d/%m/%Y")

        cliente = Cliente.query.filter_by(
            id_contacto=id_contacto,
        ).first()

        extracto = Extracto.query.filter_by(
            id_cliente=cliente.id, periodo=formatted_period, tipo_extracto=tipo_extracto
        ).first()

        if not cliente or not extracto:
            raise ValueError("Cliente, periodo o tipo de extracto no encontrado.")

        productos = Producto.query.filter_by(
            id_cliente=cliente.id, periodo=formatted_period
        ).all()

        extracto.visitado = True
        extracto.veces_visitado += 1
        extracto.ultima_visita = datetime.now()

        db.session.commit()

        template_name = get_template_for_extracto(tipo_extracto)
        html = render_template(
            template_name,
            cliente=cliente,
            registros=productos,
            periodo=periodo,
            generation_date=generation_date,
        )
        pdf_data = HTML(string=html).write_pdf()
        encrypted_pdf_data = encrypt_pdf(pdf_data, str(cliente.id))

        response = Response(encrypted_pdf_data, content_type="application/pdf")
        response.headers[
            "Content-Disposition"
        ] = f"inline; filename={tipo_extracto}_{cliente.id_contacto}_{formatted_period}.pdf"
        return response

    except ValueError as ve:
        return jsonify(error=str(ve)), 404

    except Exception as e:
        print(e)
        logging.exception(e)
        return (
            jsonify(error="Un error interno ha ocurrido. Por favor intenta más tarde."),
            500,
        )


def get_template_for_extracto(tipo_extracto):
    if tipo_extracto == "estado-cuenta":
        return "extracto_estado_cuenta.html"
    return "extracto_mora.html"


@extractos.route("/subir-extracto", methods=["GET"])
def upload_template():
    return render_template("upload_file.html")


@extractos.route("/upload-extracto", methods=["POST"])
def upload_excel():
    logging.info("Starting file upload process...")

    # Validaciones del archivo
    file = request.files.get("file")
    if not file:
        logging.error("No file part in request")
        return jsonify({"error": "No file part"}), 400
    if file.filename == "":
        logging.error("No file selected")
        return jsonify({"error": "No selected file"}), 400
    if not allowed_file(file.filename):
        logging.error("File type not allowed")
        return jsonify({"error": "File type not allowed"}), 400

    try:
        # Extraer la extensión del archivo original
        file_extension = os.path.splitext(file.filename)[1]

        # Renombrar el archivo con el timestamp actual
        timestamp = datetime.now().strftime("%d-%m-%Y-%H%M%S")
        new_filename = f"extracto-{timestamp}{file_extension}"

        filepath = os.path.join(current_app.config["UPLOAD_FOLDER"], new_filename)
        file.save(filepath)
        process_excel(filepath)
        return (
            jsonify({"message": "Datos importados y correos enviados exitosamente!"}),
            200,
        )

    except Exception as e:
        logging.error(f"Error while processing file: {e}")
        return jsonify({"error": f"Error processing file: {str(e)}"}), 500


def process_excel(filepath):
    errors = []
    upload_id = f"{str(uuid.uuid4())[:8]}-{datetime.now().strftime('%Y%m%d%H%M%S')}"

    try:
        workbook = load_workbook(filepath)
        sheet = workbook.active

        for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
            try:
                logging.info(f"Processing row {idx}: {row}")
                validate_row(row, idx)
                process_data(row, upload_id)
            except Exception as e:
                error_message = f"Error in row {idx}: {str(e)}"
                errors.append(error_message)
                logging.error(error_message)

        try:
            db.session.commit()
            send_extractos(upload_id)
        except Exception as e:
            error_message = f"Error processing file: {str(e)}"
            errors.append(error_message)
            logging.error(error_message)

    except Exception as e:
        error_message = f"Error opening/reading the Excel file: {str(e)}"
        errors.append(error_message)
        logging.error(error_message)
    finally:
        workbook.close()

    error_filepath = f"error_log_{upload_id}.csv"

    # Write the errors to a file if there are any.
    if errors:
        with open(error_filepath, "w", newline="") as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow(["Errors"])
            for error in errors:
                writer.writerow([error])

        return error_filepath
    return "Processing completed!"


def process_data(row, upload_id):
    create_or_update_cliente(row)

    tipo_extracto = row[16]

    common_data = {
        "id_cliente": row[1],
        "periodo": row[9],
        "numero_producto": row[10],
    }

    if tipo_extracto == "extracto-mora":
        additional_data = {
            "dia_pago": row[11],
            "valor_cuota_mensual": row[12],
            "valor_mora": row[13],
            "dias_mora": row[14],
            "siguiente_pago": row[15],
        }

    elif tipo_extracto == "estado-cuenta":
        additional_data = {
            "numero_obligaciones": row[11],
            "saldo_total": row[12],
            "meses_castigo": row[13],
            "fecha_corte": row[14],
        }

    else:
        raise ValueError(f"Tipo de extracto no válido: {tipo_extracto}")

    product_data = {**common_data, **additional_data}
    add_to_db(Producto, product_data)

    add_to_db(
        Extracto,
        {
            "id_cliente": row[1],
            "id_contacto": row[0],
            "periodo": row[9],
            "codigo_de_cargue": upload_id,
            "tipo_extracto": row[16],
        },
    )


def create_or_update_cliente(row):
    cliente = Cliente.query.filter_by(id_contacto=row[0]).first()
    data = {
        "id": row[1],
        "id_contacto": row[0],
        "nombre_titular": row[2],
        "telefono_1": row[3],
        "telefono_2": row[4],
        "telefono_3": row[5],
        "mail_1": row[6],
        "mail_2": row[7],
        "mail_3": row[8],
        "periodo": row[9],
    }
    if not cliente:
        add_to_db(Cliente, data)
    else:
        update_existing_cliente(cliente, data)


def update_existing_cliente(cliente, data):
    for key, value in data.items():
        setattr(cliente, key, value)
    db.session.commit()


def add_to_db(model, data):
    instance = model(**data)
    db.session.add(instance)


def handle_db_exception(e):
    logging.exception(e)
    db.session.rollback()

    error_messages = {
        "cliente": f"Un error ha ocurrido al cargar el cliente {e}",
        "producto": f"Un error ha ocurrido al cargar el producto {e}",
        "extracto": f"Un error ha ocurrido al cargar el extracto {e}",
    }

    for key, message in error_messages.items():
        if key in str(e).lower():
            raise Exception(message)

    raise Exception(f"Un error ha ocurrido al enviar los correos {e}")


@extractos.route("/send-extractos/<codigo_de_cargue>", methods=["POST"])
def send_extractos(codigo_de_cargue):
    try:
        extractos = Extracto.query.filter_by(codigo_de_cargue=codigo_de_cargue).all()

        for extracto in extractos:
            cliente = Cliente.query.filter_by(id=extracto.id_cliente).first()
            tipo_extracto = extracto.tipo_extracto

            if cliente is None:
                raise Exception(
                    f"No se encontró el cliente con id {extracto.id_cliente}"
                )

            status = send_extracto_email(cliente, tipo_extracto)

            if status != 200:
                raise Exception(
                    f"Error al enviar correo a {cliente.nombre_titular} Correo:{get_primary_email(cliente)} - Status: {status}"
                )

        return jsonify({"message": "Emails enviados exitosamente!"})

    except Exception as e:
        return jsonify({"error": str(e)}), 500


def send_extracto_email(cliente, tipo_extracto, max_retries=3):
    all_headers = {
        "Authorization": Config.MAIL_API_KEY,
    }

    periodo = date_to_period(cliente.periodo)
    extracto_url = f"{Config.APP_BASE_URL}/extractos/{tipo_extracto}/{periodo}/{cliente.id_contacto}.pdf"
    files = ["xd"]

    SENDER_EMAIL = Config.SENDER_EMAIL
    RECIPIENT_EMAIL = get_primary_email(cliente)
    EMAIL_SUBJECT = f" Extracto de tus productos QNT - {periodo.split('-')[0]}"
    BULK_ID = f"{tipo_extracto}-{periodo.split('-')[0]}"
    MESSAGE_ID = f"{tipo_extracto}-{periodo.split('-')[0]}-{cliente.id_contacto}"

    # Render the HTML template with the actual values
    email_html = render_template(
        "email_template.html",
        nombre_titular=cliente.nombre_titular,
        extracto_url=extracto_url,
    )

    form_data = {
        "from": SENDER_EMAIL,
        "replyTo": Config.EMAIL_REPLYTO,
        "to": RECIPIENT_EMAIL,
        "subject": EMAIL_SUBJECT,
        "bulkId": BULK_ID,
        "messageId": MESSAGE_ID,
        "html": email_html,
    }

    for attempt in range(max_retries):
        response = requests.post(
            Config.MAIL_BASE_URL + "/email/3/send",
            data=form_data,
            files=files,
            headers=all_headers,
        )

        # Si el código de respuesta es 200, se envió correctamente.
        if response.status_code == 200:
            return response.status_code

        # Si recibimos un 400, esperamos y reintentamos.
        elif response.status_code == 400:
            time.sleep(5)  # Espera 5 segundos antes de reintentar.
            continue

        # Si es otro código de error, levantamos una excepción.
        else:
            raise Exception(
                f"Error al enviar correo a {cliente.nombre_titular} Correo:{get_primary_email(cliente)} - Status: {response.status_code}"
            )

    # Si hemos agotado todos los reintento y aún así ha fallado, levantamos una excepción.
    raise Exception(
        f"Error después de {max_retries} intentos. No se pudo enviar el correo a {cliente.nombre_titular} Correo:{get_primary_email(cliente)} - Último Status: {response.status_code}"
    )
